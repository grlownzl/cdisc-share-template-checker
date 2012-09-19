#!/usr/bin/env python

"""
Checks a Content Template Sheet
"""

import os
import sys
import glob
import time
import re
import inspect
import json

# XML Spreadsheet format
import openpyxl
from openpyxl.workbook import Workbook
from openpyxl.style import Color, Fill
from openpyxl.cell import Cell
from openpyxl.cell import get_column_letter

"""
Extract content from an unknown state
"""

BRIDG_VERSION = "3.0.3"

# General Dictionary of codes to 
MAPPING_CODES = {'Mapping to BRIDG Defined Class' : 'BRIDG Defined Class C-Code',
                 'Mapping to BRIDG Defined Class Attribute' : 'BRIDG Defined Class Attribute C-Code',
                 'Mapping to BRIDG Performed Class' : 	'BRIDG Performed Class C-Code',
                 'Mapping to BRIDG Performed Class Attribute' : 'BRIDG Performed Class Attribute C-Code',
                 'Mapping to BRIDG Non-defined/Non-performed Class' : 'BRIDG Non-defined/Non-performed Class C-Code',
                 'Mapping to BRIDG Non-defined/Non-performed Class Attribute' : 'BRIDG Non-defined/Non-performed Class Attribute C-Code',
                 'Variable Name' : 'Variable Name C-Code',
                 'ISO 21090 Datatype' :	'ISO 21090 Datatype C-Code',
                 'ISO 21090 Datatype Constraint' : 'ISO 21090 Datatype Constraint C-Code',
                 'Description of Observation, ObservationResult or Activity or Relationship - CODED VALUES' : 'Description of Observation, ObservationResult or Activity or Relationship - CODED VALUES C-Code'}

MAPPING_ORDER = ['Variable Name', 
                'Mapping to BRIDG Defined Class', 
                'Mapping to BRIDG Defined Class Attribute',
                'Mapping to BRIDG Performed Class', 
                'Mapping to BRIDG Performed Class Attribute',
                'Mapping to BRIDG Non-defined/Non-performed Class', 
                'Mapping to BRIDG Non-defined/Non-performed Class Attribute',
                'ISO 21090 Datatype', 
                'ISO 21090 Datatype Constraint', 
                'Description of Observation, ObservationResult or Activity or Relationship - CODED VALUES']

COLUMNS = {"GENERIC" : [u'Variable Name',
                        u'Variable Name C-Code',
                        u'Variable Label',
                        u'SHARE Generic Definition',
                        u'SDTM IG 3.1.2',
                        u'SEND 3.0',
                        u'CDASH V1.1',
                        u'CDASH V1.1 Conceptual Datatype',
                        u'SDTM IG 3.1.2 Datatype',
                        u'Codelist Master',
                        u'Set of Valid Values',
                        u'Assigned Value',
                        u'Mapping to BRIDG Defined Class',
                        u'Mapping to BRIDG Defined Class Attribute',
                        u'BRIDG Defined Class C-Code',
                        u'BRIDG Defined Class Attribute C-Code',
                        u'Mapping to BRIDG Performed Class',
                        u'Mapping to BRIDG Performed Class Attribute',
                        u'BRIDG Performed Class C-Code',
                        u'BRIDG Performed Class Attribute C-Code',
                        u'Mapping to BRIDG Non-defined/Non-performed Class',
                        u'Mapping to BRIDG Non-defined/Non-performed Class Attribute',
                        u'BRIDG Non-defined/Non-performed Class C-Code',
                        u'BRIDG Non-defined/Non-performed Class Attribute C-Code',
                        u'Mapping to BRIDG Planned Class',
                        u'Mapping to BRIDG Planned Class Attribute',
                        u'BRIDG Planned Class C-Code',
                        u'BRIDG Planned Class Attribute C-Code',
                        u'ISO 21090 Datatype', u'ISO 21090 Datatype C-Code',
                        u'ISO 21090 Datatype Component',
                        u'AsCollectedIndicator',
                        u'Observation, ObservationResult, Activity, Relationship',
                        u'Description of Observation, ObservationResult or Activity or Relationship - CODED VALUES',
                        u'Description of Observation, ObservationResult or Activity or Relationship - C-Codes',
                        u'Description of Observation, ObservationResult or Activity or Relationship - NON-CODED VALUES',
                        u'NOTES'],
        "TEMPLATE" : [u'Variable Name',
                      u'Variable Label',
                      u'Codelist Master',
                      u'Set of Valid Values',
                      u'Assigned Value',
                      u'Null Flavors',
                      u'Boolean Mapping',
                      u'ISO 21090 Datatype Constraint',
                      u'ISO 21090 Datatype Constraint C-Code',
                      u'ISO 21090 Datatype Constraint Attribute',
                      u'Observation or Activity']
                      }

"""
Rules per the Team
Within header BRIDG version and Domain are mandatory. See below for column rules:
A - populated, no blanks
B - populated, no blanks
C - populated, no blanks
D - populated, no blanks
E - populated, no blanks
F - should be blank
G - populated, no blanks
H - If G = Y, should be populated
I - If E = Y, should be populated
J - okay to be blank
K - okay to be blank
L - does not have to populated
M - should be populated with NA, if not the C-code should be present
N - should be populated with NA or data, no blanks
O - AB may have data or be blank depending on BRIDG entries
If AC is populated, AE should contain DT component
AG, AH, AJ should all be populated, if not with a real value, then NA
AI - C-code should be present (linked with AH)
AK - optional, okay to be blank
o Should be one set of planned and defined for each concept, but some are
implementation specific like SEQ
o Can one variable have both defined and performed? Per Diane, answer = yes
o If any BRIDG class is populated than the DT and C-code fields should be populated
"""

# Columns that must be set
MUSTSET = {u'Variable Name' : True,
           u'Variable Name C-Code' : True,
           u'Variable Label' : True,
           u'SHARE Generic Definition' : True,
           u'SDTM IG 3.1.2' : True,
           u'CDASH V1.1' : True,
           u'CDASH V1.1 Conceptual Datatype' : {"Y" : u'CDASH V1.1'},
           u'SDTM IG 3.1.2 Datatype' : {"Y" : 'SDTM IG 3.1.2'},
           u'ISO 21090 Datatype' : {"SET" :[u'Mapping to BRIDG Defined Class',
                                            u'Mapping to BRIDG Performed Class',
                                            u'Mapping to BRIDG Non-defined/Non-performed Class',
                                            u'Mapping to BRIDG Planned Class']},
        u'ISO 21090 Datatype Component' : {"SET" : u'ISO 21090 Datatype'},
           u'Observation, ObservationResult, Activity, Relationship' : True,
           }

# Set of columns that must not be set
MUSTNOTSET = {"SEND 3.0" : True}

# Set of columns that must be set or NA
MUSTVALORNA = {u'Mapping to BRIDG Defined Class' : True,
               u'Mapping to BRIDG Defined Class Attribute' : True,
               u'Mapping to BRIDG Performed Class' : True,
               u'Mapping to BRIDG Performed Class Attribute' : True,
               u'Mapping to BRIDG Non-defined/Non-performed Class' : True,
               u'Mapping to BRIDG Non-defined/Non-performed Class Attribute' : True,
               u'Mapping to BRIDG Planned Class' : True,
               u'Mapping to BRIDG Planned Class Attribute' : True,
               u'Observation, ObservationResult, Activity, Relationship' : True,
               u'Description of Observation, ObservationResult or Activity or Relationship - CODED VALUES' : True,
               u'Description of Observation, ObservationResult or Activity or Relationship - NON-CODED VALUES' : True
               }

MUTEX = [( u'Description of Observation, ObservationResult or Activity or Relationship - CODED VALUES',
           u'Description of Observation, ObservationResult or Activity or Relationship - NON-CODED VALUES')]

# Reverse Dependencies
REVERSE_DEPS = {u'CDASH V1.1 Conceptual Datatype' : { u'CDASH V1.1' : "N"},
                u'SDTM IG 3.1.2 Datatype' : {'SDTM IG 3.1.2' : "N"},
                }

def columnify(set_of_columns):
  """
  return a dict with the column indicies (per Excel)
  """
  mapped = {}
  for (idx, col) in enumerate(set_of_columns):
    prefix = {0 : ''}.get(idx/26, chr(65 + idx/26 - 1))
    mapped["%s%s" % (prefix, chr(65 + idx % 26))] = col
  return mapped

def si(content):
  """
  If it gets a None, return '', else return cleaned string
  """
  import types
  if isinstance(content, openpyxl.cell.Cell):
    if isinstance(content.value, types.NoneType):
        return unicode('')
    else:
        return unicode(content.value).strip()
  elif isinstance(content, xlrd.sheet.Cell):
    if content.ctype == xlrd.sheet.XL_CELL_EMPTY:
        return unicode('')
    else:
        return unicode(content.value).strip()
  return unicode(content).strip()

class ContentSheetChecker(object):

  def __init__(self):
    self.templates = []
    self.template_vars = {}
    self.exceptions = {}
    self.template = ""
    self.sheet = ""
    self._rules = []
    self._get_rules()
  
  @property
  def has_issues(self):
    return len(self.exceptions) != 0
          
  def log(self, field, column, message):
    # Log an exception
    self.exceptions.setdefault(self.template, []).append([self.sheet, field, column, message])

  def as_dict(self):
    """
    serialise the issues as a list of dicts
    """
    serialised = []
    for (template, issues) in self.exceptions.iteritems():
      for issue in issues:
        serialised.append({'template' : template,
                            'sheet' : issue[0],
                            'field' : issue[1],
                            'column' : issue[2],
                            'message' : issue[3]})
    return serialised
    
  def as_json(self):
    """
    return the issues as a json dump
    """
    return json.dumps(self.as_dict())
      
  def report(self, exceptions=[]):
    if len(self.exceptions) > 0:
      report = Workbook()
      for (contentbook, exceptions) in self.exceptions.iteritems():
        fname = os.path.basename(contentbook)
        template = os.path.splitext(fname)[0]
        if "Sheet" in report.get_sheet_names():
          ws = report.get_sheet_by_name("Sheet")
        else:
          ws = report.create_sheet()
        t = re.compile("([A-Z][A-Z\s\-]+) Template")
        title = t.search(template).groups()[0]
        ws.title = title
        sheet = report.get_sheet_by_name(title)
        for (idx, colname) in enumerate(["Sheet", "Field", "Column", "Issue"], 1):
          col_letter = get_column_letter(idx)
          _cell = sheet.cell("%s1" % col_letter)
          _cell.style.font.name = 'Arial'
          _cell.style.font.size = 12
          _cell.style.font.bold = True
          _cell.style.alignment.vertical = openpyxl.style.Alignment.VERTICAL_TOP
          _cell.style.alignment.wrap_text = True
          _cell.value = colname
        for (row_idx, exc_item) in enumerate(exceptions, 2):
          for (col_idx, field) in enumerate(exc_item, 1):
            _cell = sheet.cell("%s%s" % (get_column_letter(col_idx), row_idx))
            _cell.value = field
        for column in range(1, 5):
          if column in [1, 4]:
            width = "30"
          elif column in [2]:
            width = 15
          elif column in [3]:
            width = 75
          sheet.column_dimensions[get_column_letter(column)].width = width
      report.save("Content_Template_Check_%s.xlsx" % time.strftime("%Y-%m-%d"))
                    
    else:
      print "Nothing to report"

  def _get_rules(self):
    """
    Load the rules
    """
    if self._rules == []:
      _all_funcs = [x for x in inspect.getmembers(self, inspect.ismethod) if x[0].startswith("_run_")]
      for (funcname, func) in _all_funcs:
        self._rules.append(func)

  @property
  def rules(self):
    """
    return the rules
    """
    return self._rules
    
  @property
  def rule_specifications(self):
    """
    List the rules in action
    """
    return [x.__doc__ for x in self._rules]
    
  def load_from_mem(self, memobj):
    self.templates.append(memobj.filename)
    self.template = memobj.filename
    try:
      workbook = openpyxl.reader.excel.load_workbook(memobj.file)
    except Exception, e:
      import traceback
      print 'Failed to open %s : %s' % (memobj.filename, e)
      traceback.print_tb(sys.exc_info()[2])
      return
    for sheet_name in workbook.get_sheet_names():
      sheet = workbook.get_sheet_by_name(sheet_name)
      if sheet.cell("A1").value.upper() not in ["BRIDG VERSION", "CONCEPT"]:
        # Only look at those with a BRIDG Version top left
        continue
      if sheet.cell("C1").value and sheet.cell("C1").value.upper() == "WIP":
        self.log(sheet_name, "ALL", "ALL", "Sheet has been marked as Work in Progress and has not been scanned")
        continue
      self.sheet = sheet_name
      self.run_checks(sheet)

  def load_from_file(self, contentbook):
    self.templates.append(contentbook)
    self.template = contentbook
    try:
      workbook = openpyxl.reader.excel.load_workbook(contentbook)
    except Exception, e:
      import traceback
      print 'Failed to open %s : %s' % (contentbook, e)
      traceback.print_tb(sys.exc_info()[2])
      return
    for sheet_name in workbook.get_sheet_names():
      sheet = workbook.get_sheet_by_name(sheet_name)
      if sheet.cell("A1").value.upper() not in ["BRIDG VERSION", "CONCEPT"]:
        # Only look at those with a BRIDG Version top left
        continue
      self.sheet = sheet_name
      self.run_checks(sheet)

  def run_checks(self, sheet):
    """
    Run all checks 
    """
    COLS = []
    for (offset, row) in enumerate(sheet.rows, 1):
      if si(row[0]) == "":
        continue
      elif si(row[0]).upper() == "BRIDG VERSION":
        if not (si(row[1]) == BRIDG_VERSION or si(row[2]) == BRIDG_VERSION):
            self.log("", si(row[0]), "BRIDG Version not set or not equal to %s" % BRIDG_VERSION)
      elif si(row[0]) == "Domain":
        if si(row[1]) == "":
            self.log("", si(row[0]), "Domain not set")
      elif si(row[0]).upper() == "VARIABLE NAME":
        if 'GENERIC' in self.sheet.upper():
          COLS = COLUMNS.get('GENERIC')
        else:
          COLS = COLUMNS.get('TEMPLATE')
        if len(row) != len(COLS):
          _COLS = [si(x) for x in row]
          extra = ','.join(set(_COLS) - set(COLS))
          missing = ','.join(set(COLS) - set(_COLS))
          self.log("", "HEADINGS", "Number of columns doesn't meet expectations: extras '%s' - missing '%s'" % (extra, missing))
          COLS = _COLS
        break
    # yuck
    for (row_idx, contentrow) in enumerate(sheet.rows[offset:], offset):
      if int(row_idx) < int(offset):
        continue
      mapped = dict(zip(COLS, [si(x) for x in contentrow]))
      if mapped.get('Variable Name') == "":
        # skip blanks
        continue
      if 'GENERIC' in self.sheet.upper():
        # populate template vars on Generic Tab
        self.template_vars.setdefault(self.template, []).append(mapped.get('Variable Name'))
      for rule in self.rules:
        rule.__call__(mapped)

  def _run_reverse_dependencies(self, row):
    """
    Check that a dependent column is missing if 'depends on' is not set, or set to a value excluding setting
    """
    for (col, dependencies) in REVERSE_DEPS.iteritems():
      for (col_dep, dep_val) in dependencies.iteritems():
        if row.get(col_dep) == dep_val:
          if row.get(col) != "": 
            self.log(row.get('Variable Name'),
                  col,
                  "Should not be set, as is a dependent variable")

  def _run_check_bridg_is_set(self, row):
    """
    Check that at least one BRIDG attribute is set (except for DOMAIN)
    """
    # isolate BRIDG columns
    cols = [x for x in row.keys() if x.startswith("Mapping to BRIDG")]
    if len(cols) == 0:
      return
    set_values = [row.get(x) for x in cols if row.get(x) != ""]
    if len(set_values) == 0:
      self.log(row.get("Variable Name"),
                      "BRIDG Mappings",
                      "No BRIDG Mapping currently assigned to %s" % row.get('Variable Name'))
        
      
  def _run_check_copying_from_generic(self, row):
    """
    Check that all fields in the Concept Tabs are represented in the Generic Tab
    """
    if not 'GENERIC' in self.sheet.upper():
      try:
        if not row.get('Variable Name') in self.template_vars.get(self.template):
          self.log(row.get("Variable Name"),
                   "Variable name",
              "Variable %s is in a Concept Tab, but not in the Generic Tab" % row.get('Variable Name'))
      except TypeError:
        print "Template Vars: %s" % self.template_vars.get(self.template)
    
  def _run_check_bridg_attributes_classes(self, row):
    """
    Check that the BRIDG classes/attributes are valid values
    """
    pass
                  
  def _run_set_check(self, row):
    """
    Checks that a column is populated when it should be
    """
    for (column, content) in row.iteritems():
      if MUSTSET.get(column):
        # column flagged as needing setting
        if content == "":
          # nothing set
          if MUSTSET.get(column) == True:
            self.log(row.get('Variable Name'),
                     column,
                     "Column must be set")
          else:
            # check the dependency
            for (depval, dep) in MUSTSET.get(column, {}).iteritems():
              # Dep is a dependent column, and deplval is a value that should be set
              if isinstance(dep, str):
                if row.get(dep) == depval:
                  self.log(row.get('Variable Name'),
                           column,
                      "Column must be set but is not - based on dependency of '%s' having value %s" % (dep, depval))
                elif (depval == "SET" and row.get(dep) != ""):
                  self.log(row.get('Variable Name'),
                           column,
                      "Column must be set but is not - based on dependency of '%s' having value %s" % (dep, depval))
              elif isinstance(dep, list):
                if depval == "SET":
                  assigned = [row.get(x) for x in dep]
                  if assigned.count("") == len(assigned):
                    self.log(row.get('Variable Name'),
                             column,
                             "Column must be set but is not - based on dependency of '%s' having value %s" % (','.join(dep), depval) )
                                        
  def _run_set_or_na_check(self, row):
    """
    Check that field is set or NA
    """
    for (column, content) in row.iteritems():
      if MUSTVALORNA.get(column, False) == True:
        if content == "":
          self.log(row.get('Variable Name'),
                   column,
              "Column must be set to value or na")
        
        
  def _run_not_set_check(self, row):
    """
    Check that a column is not set
    """
    for (column, content) in row.iteritems():
      if MUSTNOTSET.get(column, False) == True:
        if content != "":
          self.log(row.get('Variable Name'),
                   column,
              "Column is set when it shouldn't be")
                
  def _run_check_coding_columns(self, row):
    """
    Check that all codable elements have been coded
    """
    for (target, targetted) in MAPPING_CODES.iteritems():
      if not row.get(target) in ["", "na"] and row.get(targetted) == "":
        self.log(row.get('Variable Name'),
            targetted,
            "Expected Coding is missing")
    
if __name__ == "__main__":
  import optparse
  parser = optparse.OptionParser()
  parser.add_option("-p", "--path", dest="prefix", default=os.getcwd(), action="store")
  (opts, args) = parser.parse_args()
  checker = ContentSheetChecker()
  for candidate in glob.glob(os.path.join(opts.prefix, "*.xls")) + glob.glob(os.path.join(opts.prefix, "*.xlsx")):
    if '~' in candidate:
      # skip temp files
      continue
    if candidate.endswith("Template.xls") or candidate.endswith("Template.xlsx"):
      print "Checking %s" % candidate
      checker.load_from_file(candidate)
  print checker.as_dict()
  checker.report()
        
    
